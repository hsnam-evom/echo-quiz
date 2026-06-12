#!/usr/bin/env python3
"""echo_quiz_100.xlsx → questions.json 변환 스크립트.

사용법:
    python3 scripts/build-questions.py ../echo_quiz_100.xlsx questions.json

요구사항: openpyxl (pip install openpyxl)
"""
import json
import sys

import openpyxl


def main(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Echo Quiz 100"]

    # 유형 코드 → 분류명 매핑
    cat = {}
    for row in list(wb["유형 분류"].iter_rows(values_only=True))[1:]:
        if row[0]:
            cat[row[0]] = row[1]

    out = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        num, typ, q, ans, w1, w2, w3, exp = row
        out.append({
            "id": int(num),
            "type": typ,
            "category": cat.get(typ, ""),
            "question": q.strip(),
            "answer": ans.strip(),
            "distractors": [w1.strip(), w2.strip(), w3.strip()],
            "explanation": (exp or "").strip(),
        })

    with open(dst, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=0)
    print(f"wrote {len(out)} questions -> {dst}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "../echo_quiz_100.xlsx"
    dst = sys.argv[2] if len(sys.argv) > 2 else "questions.json"
    main(src, dst)
